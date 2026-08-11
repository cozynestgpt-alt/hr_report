"""
월례회의 인원보고서 자동화 스크립트
사용법: python main.py [YYYYMM]
예시:  python main.py 202605
       python main.py           ← 당월 자동 적용

입력 파일:  input/부서별인원현황_YYYYMMDD.xlsx
            컬럼 구조: 부서코드 | 부서 | 99년미만 | 계 | 평균
            input/★월례회의_인원보고_양식.xlsx
            → 중간관리·물류센터용역 수기입력 양식. 담당자가 매월 직접 갱신.
              L2(기준일)가 당월 말일과 일치할 때만 자동 반영됨.

출력 파일:  output/월례회의_인원보고_YYYYMM_생성일자.xlsx
            (동일 파일 존재 시 _v2, _v3 ... 로 자동 증가, 기존 결과물은 덮어쓰지 않음)

자동 생성 범위:
  ○ 브랜드 사업부문 일반직 (행 5~19): input 파일 부서코드 기준 자동 집계
  ○ 영업본부 백화점 일반직 (행 7):    원팀+올라운더팀 합산
  ○ 영업본부 백화점 판매직 (행 7):    위탁매장(AR2xxx) 전체 합산
  ○ 직영점 판매직 (행 21~32):          직영점별 자동 집계
  ○ 중간관리 (행 34~37):               ★양식 파일 기준일 일치 시 자동 반영, 아니면 수동 입력
  ○ 물류센터용역·평택점용역 (행 43~44): ★양식 파일 기준일 일치 시 자동 반영, 아니면 수동 입력
  △ 신규입사자명단·퇴사자명단 시트:     수동 입력 필요
"""

import sys
import os
import glob
import shutil
import subprocess
import json
import calendar
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

for _stream in (sys.stdout, sys.stderr):
    if hasattr(_stream, "reconfigure"):
        _stream.reconfigure(encoding="utf-8", errors="replace")


# ── 경로 설정 ─────────────────────────────────────────────
BASE_DIR   = Path(__file__).parent
INPUT_DIR  = BASE_DIR / "input"
OUTPUT_DIR = BASE_DIR / "output"
TEMPLATE   = BASE_DIR / "template" / "월례회의_인원보고_템플릿.xlsx"
RECALC     = BASE_DIR / "scripts" / "recalc.py"
MANUAL_FILE = INPUT_DIR / "★월례회의_인원보고_양식.xlsx"   # 중간관리·물류센터용역 수기입력 양식 (담당자가 매월 갱신)

REPORT_SHEET = "전년대비 인원증감현황"   # 템플릿 고정 시트명

# ── 양식 파일에서 그대로 복사해오는 셀 (중간관리 행 34~37, 물류센터용역·평택점용역 43~44) ──
MANUAL_COPY_CELLS = [f"{col}{r}" for r in (34, 35, 36, 37) for col in ("E", "F", "G", "H")] + ["C43", "C44"]

COL_CODE  = "부서코드"   # 부서코드 열
COL_COUNT = "계"         # 인원수 열 (문자열로 저장됨)

# ── 구 부서코드 → 현 부서코드 매핑 ────────────────────────
#   전년도 파일에 구 코드가 남아있을 경우 현 코드로 통합
OLD_CODE_MAP = {
    "AR1300": "AR3104",   # ETC 구 코드
    "AR3101": "AR3102",   # PLD → PDD 통합
    "AR3103": "AR3105",   # PMD → MMD 통합
}

# ── 직영점으로 분류되는 AR2xxx 코드 (영업본부 백화점 집계에서 제외) ──
JIKGYEONG_AR2 = {"AR2501", "AR2683"}

# ── 보고서 행 구조 ─────────────────────────────────────────
# 형식: (엑셀행번호, 설명, [일반직 부서코드], [판매직 부서코드])
#   - 빈 리스트 []                    → 해당 구분 집계 없음 (셀 유지)
#   - "AR2_ALL_EXCEPT_JIKGYEONG"      → AR2xxx 전체 합산(직영점 제외)
#   - "SKIP"                          → 행을 건드리지 않음 (수동 입력 전용)
REPORT_ROWS = [
    # ── 브랜드 사업부문 ───────────────────────────────────────
    (5,  "임원",             ["AR1003", "AR1005", "AR1006"], []),
    (6,  "영업본부_총괄",     ["AR4100"], []),
    (7,  "영업본부_백화점",   ["AR4104", "AR4105"],                # 일반: 원팀 + 올라운더팀
                               "AR2_ALL_EXCEPT_JIKGYEONG"),        # 판매: AR2xxx 전체 (직영점 제외)
    (8,  "영업본부_SMD",      ["AR4103"], []),
    (9,  "영업본부_영업지원", ["AR4101"], []),
    (10, "온라인사업부",      ["AR1700"], []),
    (11, "상품본부_총괄",     ["AR3100"], []),
    (12, "상품본부_PDD",      ["AR3102"], []),                     # PLD(AR3101) → AR3102 로 통합
    (13, "상품본부_MMD",      ["AR3105"], []),                     # PMD(AR3103) → AR3105 로 통합
    (14, "RND_총괄",          ["AR1101"], []),
    (15, "RND",               ["AR1121", "AR1122", "AR1123"], []),
    (16, "VMD",               ["AR1117"], []),
    (17, "AMD",               ["AR1200"], []),
    (18, "CXD",               ["AR1510"], []),
    (19, "ETC",               ["AR3104"], []),                    # 구 ETC(AR1300) → AR3104
    # ── 직영점 (판매직만) ─────────────────────────────────────
    (21, "직영점_양재점",             [], ["AR1960"]),
    (22, "직영점_일산점",             [], ["AR1800"]),
    (23, "직영점_경기광주점",         [], ["AR1790"]),
    (24, "직영점_전주점",             [], ["AR1910"]),
    (25, "직영점_NC충장점",           [], ["AR1940"]),
    (26, "직영점_NC불광점",           [], ["AR1950"]),
    (27, "직영점_청주점",             [], ["AR1970"]),
    (28, "직영점_NC해운대점",         [], ["AR1980"]),
    (29, "직영점_NC일산점",           [], ["AR1981"]),
    (30, "직영점_고양점(롯데아울렛)", [], ["AR2501"]),
    (31, "직영점_가든5점_현대아울렛", [], ["AR2683"]),
    (32, "직영점_현대커넥트부산점",   [], ["AR1982"]),
    # ── 중간관리 (수동 입력) ──────────────────────────────────
    (34, "중간관리_롯데   ← 수동입력", "SKIP", "SKIP"),
    (35, "중간관리_현대   ← 수동입력", "SKIP", "SKIP"),
    (36, "중간관리_신세계 ← 수동입력", "SKIP", "SKIP"),
    (37, "중간관리_갤러리아← 수동입력", "SKIP", "SKIP"),
]


# ── 내부 함수 ─────────────────────────────────────────────
def last_day(year: int, month: int) -> date:
    return date(year, month, calendar.monthrange(year, month)[1])


def find_file(year: int, month: int) -> Path | None:
    end = last_day(year, month)
    exact = INPUT_DIR / f"부서별인원현황_{end:%Y%m%d}.xlsx"
    if exact.exists():
        return exact
    pattern = str(INPUT_DIR / f"부서별인원현황_{year}{month:02d}*.xlsx")
    cands = sorted(glob.glob(pattern))
    if cands:
        print(f"  ※ 말일자 파일 없음 → 사용: {os.path.basename(cands[-1])}")
        return Path(cands[-1])
    return None


def read_counts(filepath: Path) -> dict:
    """부서별인원현황 파일 → {부서코드: 인원수} 딕셔너리 반환"""
    wb = load_workbook(filepath, data_only=True)
    ws = wb.active

    hdr_row = col_code = col_count = None
    for r in ws.iter_rows(max_row=10):
        for c in r:
            v = str(c.value).strip() if c.value is not None else ""
            if v == COL_CODE:
                col_code = c.column
                hdr_row = c.row
            if v == COL_COUNT:
                col_count = c.column
        if hdr_row:
            break

    if hdr_row is None:
        raise ValueError(
            f"헤더 행을 찾지 못했습니다: {filepath}\n"
            f"  → COL_CODE='{COL_CODE}', COL_COUNT='{COL_COUNT}' 확인 필요"
        )

    result = {}
    for row in ws.iter_rows(min_row=hdr_row + 1, values_only=True):
        raw_code = row[col_code - 1]
        raw_count = row[col_count - 1]
        if not raw_code:
            continue
        code = OLD_CODE_MAP.get(str(raw_code).strip(), str(raw_code).strip())
        try:
            n = int(float(str(raw_count).strip()))
        except (ValueError, TypeError):
            print(f"  ⚠ 인원수 파싱 실패 → 0으로 처리: {filepath.name} 부서코드={code} 값={raw_count!r}")
            n = 0
        result[code] = result.get(code, 0) + n

    wb.close()
    return result


def read_manual_values(cy_end: date):
    """★월례회의_인원보고_양식.xlsx(중간관리·물류센터용역 수기입력 양식)에서 값을 읽어온다.
    파일이 없거나 L2(기준일)가 당월 말일과 다르면 (None, 사유)를 반환해 수동 입력으로 넘긴다."""
    if not MANUAL_FILE.exists():
        return None, f"{MANUAL_FILE.name} 없음"

    wb = load_workbook(MANUAL_FILE, data_only=True)
    if REPORT_SHEET not in wb.sheetnames:
        wb.close()
        return None, f"{MANUAL_FILE.name}에 '{REPORT_SHEET}' 시트 없음"
    ws = wb[REPORT_SHEET]

    l2 = ws["L2"].value
    l2_date = l2.date() if isinstance(l2, datetime) else l2
    if l2_date != cy_end:
        wb.close()
        return None, f"{MANUAL_FILE.name} 기준일({l2_date}) ≠ 당월 말일({cy_end}) → 파일 갱신 필요"

    values = {cell: ws[cell].value for cell in MANUAL_COPY_CELLS}
    wb.close()
    return values, None


def calc(data: dict, codes) -> int:
    """codes: 리스트 또는 특수값 처리"""
    if codes == "SKIP" or not codes:
        return 0
    if codes == "AR2_ALL_EXCEPT_JIKGYEONG":
        return sum(v for k, v in data.items()
                   if k.startswith("AR2") and k not in JIKGYEONG_AR2)
    return sum(data.get(c, 0) for c in codes)


# ── 메인 처리 ─────────────────────────────────────────────
def main():
    if len(sys.argv) >= 2:
        ym = sys.argv[1].strip()
    else:
        today = date.today()
        ym = f"{today.year}{today.month:02d}"
        print(f"  기준월 미입력 → 당월 적용: {ym}")

    if len(ym) != 6 or not ym.isdigit():
        print("오류: YYYYMM 형식이어야 합니다 (예: 202605)")
        sys.exit(1)

    cy_y, cy_m = int(ym[:4]), int(ym[4:])
    py_y, py_m = cy_y - 1, cy_m
    cy_end = last_day(cy_y, cy_m)
    py_end = last_day(py_y, py_m)

    print(f"\n{'='*65}")
    print(f"  보고 대상: {cy_y}년 {cy_m:02d}월    전년동월: {py_y}년 {py_m:02d}월")
    print(f"  당년 말일: {cy_end}    전년 말일: {py_end}")
    print(f"{'='*65}\n")

    # 파일 탐색
    print("[1/4] 파일 탐색...")
    cy_file = find_file(cy_y, cy_m)
    py_file = find_file(py_y, py_m)

    if not cy_file:
        print(f"\n  오류: 당월 파일 없음\n  → {INPUT_DIR}\\부서별인원현황_{cy_end:%Y%m%d}.xlsx")
        sys.exit(1)
    if not py_file:
        print(f"\n  오류: 전년동월 파일 없음\n  → {INPUT_DIR}\\부서별인원현황_{py_end:%Y%m%d}.xlsx")
        sys.exit(1)

    print(f"  당년: {cy_file.name}")
    print(f"  전년: {py_file.name}")

    # 인원 읽기
    print("\n[2/4] 인원 데이터 읽기...")
    cy = read_counts(cy_file)
    py = read_counts(py_file)

    # 미매핑 부서 코드 경고
    mapped_codes = set(OLD_CODE_MAP.keys())
    for _, _, gc, sc in REPORT_ROWS:
        if isinstance(gc, list):
            mapped_codes.update(gc)
        if isinstance(sc, list):
            mapped_codes.update(sc)
    mapped_codes.update(JIKGYEONG_AR2)
    unk = set()
    for code in (set(cy) | set(py)):
        if code not in mapped_codes and not (code.startswith("AR2") and code not in JIKGYEONG_AR2):
            unk.add(code)
    if unk:
        print(f"\n  ⚠ 보고서에 매핑되지 않은 부서코드 (REPORT_ROWS 에 추가 검토 필요):")
        for u in sorted(unk):
            print(f"    {u}  전년:{py.get(u, 0):>2}명  당년:{cy.get(u, 0):>2}명")

    # 양식 복사
    print("\n[3/4] 보고서 생성...")
    if not TEMPLATE.exists():
        print(f"\n  오류: 템플릿 파일 없음\n  → {TEMPLATE}")
        sys.exit(1)

    OUTPUT_DIR.mkdir(exist_ok=True)
    today_str = date.today().strftime("%Y%m%d")
    base_path = OUTPUT_DIR / f"월례회의_인원보고_{ym}_{today_str}.xlsx"
    if not base_path.exists():
        output_path = base_path
    else:
        version = 2
        while True:
            output_path = OUTPUT_DIR / f"월례회의_인원보고_{ym}_{today_str}_v{version}.xlsx"
            if not output_path.exists():
                break
            version += 1
        print(f"  ※ 동일 파일 존재 → 새 파일명: {output_path.name}")
    shutil.copy2(TEMPLATE, output_path)

    wb = load_workbook(output_path)
    ws = wb[REPORT_SHEET]

    ws["L2"] = datetime(cy_y, cy_m, cy_end.day)

    # 중간관리·물류센터용역 양식 파일 확인 (기준일 일치 시에만 자동 반영)
    manual_values, manual_skip_reason = read_manual_values(cy_end)
    if manual_values:
        print(f"  ✅ {MANUAL_FILE.name} 기준일 확인 완료 → 중간관리·물류센터용역 자동 반영")
    else:
        print(f"  ⚠️  중간관리·물류센터용역 자동 반영 건너뜀 ({manual_skip_reason})")

    # 값 채우기
    print(f"\n  {'행':>3}  {'항목':<30}  {'전년일반':>7} {'당년일반':>7}  {'전년판매':>7} {'당년판매':>7}")
    print(f"  {'─'*3}  {'─'*30}  {'─'*7} {'─'*7}  {'─'*7} {'─'*7}")

    for row_no, label, g_codes, s_codes in REPORT_ROWS:
        if g_codes == "SKIP" and s_codes == "SKIP":
            if manual_values:
                vals = {c: manual_values[f"{c}{row_no}"] for c in ("E", "F", "G", "H")}
                for c, v in vals.items():
                    ws[f"{c}{row_no}"] = v
                print(f"  {row_no:>3}  {label:<30}  {vals['E'] or 0:>7} {vals['G'] or 0:>7}  {vals['F'] or 0:>7} {vals['H'] or 0:>7}")
            else:
                print(f"  {row_no:>3}  {label:<30}  {'(수동)':>7} {'(수동)':>7}  {'(수동)':>7} {'(수동)':>7}")
            continue

        py_g = calc(py, g_codes)
        cy_g = calc(cy, g_codes)
        py_s = calc(py, s_codes)
        cy_s = calc(cy, s_codes)

        if g_codes and g_codes != "SKIP":
            ws[f"E{row_no}"] = py_g or None
            ws[f"G{row_no}"] = cy_g or None
        if s_codes and s_codes != "SKIP":
            ws[f"F{row_no}"] = py_s or None
            ws[f"H{row_no}"] = cy_s or None

        print(f"  {row_no:>3}  {label:<30}  {py_g:>7} {cy_g:>7}  {py_s:>7} {cy_s:>7}")

    if manual_values:
        ws["C43"] = manual_values["C43"]
        ws["C44"] = manual_values["C44"]
        print(f"  {'43':>3}  {'물류센터용역':<30}  {manual_values['C43']!s:>7}")
        print(f"  {'44':>3}  {'평택점용역':<30}  {manual_values['C44']!s:>7}")

    wb.save(output_path)
    wb.close()

    # ── 수식 재계산 (LibreOffice, 설치되어 있는 경우에만) ──
    print("\n[4/4] 수식 재계산 중...")
    result = subprocess.run(
        [sys.executable, str(RECALC), str(output_path), "60"],
        capture_output=True, text=True
    )
    recalc_info = None
    if result.stdout:
        try:
            recalc_info = json.loads(result.stdout)
        except json.JSONDecodeError:
            recalc_info = None

    if recalc_info is None:
        print("  ⚠️  수식 재계산을 건너뜁니다 (LibreOffice 미설치 또는 실행 실패).")
        print("     엑셀에서 파일을 열면 수식은 정상적으로 자동 계산됩니다.")
        if result.stderr:
            print(f"     상세: {result.stderr.strip().splitlines()[-1]}")
    elif recalc_info.get("status") == "errors_found":
        print(f"  ⚠️  수식 오류 {recalc_info['total_errors']}건: {recalc_info.get('error_summary', {})}")
    elif "error" in recalc_info:
        print(f"  ⚠️  수식 재계산 실패: {recalc_info['error']}")
    else:
        print(f"  ✅ 수식 재계산 완료 (수식 {recalc_info.get('total_formulas', '?')}개)")

    print(f"\n{'='*65}")
    print(f"  ✅ 완료!  → {output_path}")
    print(f"{'='*65}\n")
    print("  ★ Excel 열기 후 수동 입력 필요 항목:")
    if not manual_values:
        print(f"    · 행 34~37: 중간관리 (롯데/현대/신세계/갤러리아) 전년·당년 인원")
        print(f"    · 행 43~44: 물류센터용역·평택점용역 인원")
        print(f"      → input/{MANUAL_FILE.name} 을(를) 당월 기준으로 갱신 후 재실행하면 자동 반영됩니다 ({manual_skip_reason})")
    print("    · 신규입사자명단·퇴사자명단 시트")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"\n[오류 발생]\n  {type(e).__name__}: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
